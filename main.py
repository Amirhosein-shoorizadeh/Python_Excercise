import inspect
import json
import random
from collections import defaultdict
import openpyxl
import json
import csv
import datetime
import time


# 1. Calculate the area of a circle with radius 7
# 2. Convert temperature: 100Fahrenheit → Celsius
#    Formula: C = (F - 32) * 5/9
# 3. Compute compound interest:
#    principal = 1000, rate = 0.05, years = 10
#    Formula: A = P * (1 + r) ** n
# 4. Find the hypotenuse of a right triangle
#    with legs a=3, b=4 (use **0.5 for square root)
# Print each result formatted to 2 decimal places
# Expected: area = 153.94, temp = 37.78, ...

def func_1():
    print(
        "///////////////////////////////////////// " + inspect.currentframe().f_code.co_name + " ///////////////////////////////////////////")
    r = 7
    print("1. : \n {:.2f}".format(r * 3.14))
    f = 100
    print("2. : \n {:.2f}".format((f - 32) * 5 / 9))
    p = 1000
    rate = 0.05
    y = 10
    print("3. : \n {:.2f}".format(p * (1 + rate) ** y))
    leg_a = 3
    leg_b = 4
    print("4. : \n {:.2f}".format((leg_a ** 2 + leg_b ** 2) ** 0.5))


full_name = "python programming language"
# 1. Capitalize each word → "Python Programming Language"
# 2. Count how many times the letter 'p' appears (case-insensitive)
# 3. Replace 'language' with 'course'
# 4. Reverse the entire string
# 5. Check if it's a palindrome
sentence = "A man a plan a canal Panama"


# 6. Remove all spaces and check if palindrome (case-insensitive)
# Print each result with a descriptive label using f-strings

def func_2():
    print(
        "///////////////////////////////////////// " + inspect.currentframe().f_code.co_name + " ///////////////////////////////////////////")
    full_name = "python programming language"

    new_word = " ".join(word.capitalize() for word in full_name.split())
    print("1." + new_word)
    print("2. p count: {}".format(full_name.lower().count("p")))
    new_full_name = full_name.replace("language", "course")
    print("3. " + new_full_name)
    print("4. " + new_full_name[::-1])
    clean = full_name.replace(" ", "").lower()
    if clean == clean[::-1]:
        print("5. Palindrome")
    else:
        print("5. Not Palindrome")

    sentence = "A man a plan a canal panama"
    new_word = ""
    for words in sentence.split():
        new_word = new_word + words
    print("6. " + new_word)


data = "2024-07-13:ERROR:Database connection failed at line 42"


# Using only slicing and string methods (no split() for parts1-3):
# 1. Extract the date: "2024-07-13"
# 2. Extract the log level: "ERROR"
# 3. Extract the message after the second colon
# 4. Now use split(':') — how does it compare?
# 5. Extract just the number42 and convert it to int

def func_3():
    print(
        "///////////////////////////////////////// " + inspect.currentframe().f_code.co_name + " /////////////////////////////////////////")
    data = "2024-07-13:ERROR:Database connection failed at line 42"
    data_time = data[0:10]
    print("1. " + data_time)
    data_level = data[11:16]
    print("2. " + data_level)
    a = data.find(":", data.find(":") + 1)
    data_massage = data[a + 1:]
    print("3. " + data_massage)
    for word in data.split(":"):
        print("4. " + word)
    a_intiger = int(data[-2::])
    print("5. {} {}".format(a_intiger, type(a_intiger)))


temperatures = [22, 35, 18, 40, 28, 15, 33, 27, 19, 38, 25, 30]


# Without using built-in max/min/sum functions:
# 1. Find the maximum and minimum manually
# 2. Calculate the average
# 3. Count how many days were above average
# 4. Separate into two lists: hot (>= 30) and cool (< 30)
# 5. Sort the list using bubble sort (implement it yourself)
# 6. Find the median (middle value of sorted list)

def func_4():
    print(
        "///////////////////////////////////////// " + inspect.currentframe().f_code.co_name + " /////////////////////////////////////////")
    temperatures = [22, 35, 18, 40, 28, 15, 33, 27, 19, 38, 25, 30]
    sorted_temperatures = temperatures.copy()
    avg_count = 0
    i = 0
    while i < len(sorted_temperatures) - 1:
        j = 0
        while j < len(sorted_temperatures) - 1 - i:
            if sorted_temperatures[j] > sorted_temperatures[j + 1]:
                temp = sorted_temperatures[j]
                sorted_temperatures[j] = sorted_temperatures[j + 1]
                sorted_temperatures[j + 1] = temp
            j = j + 1
        avg_count = avg_count + sorted_temperatures[i]
        i = i + 1
    print("1. min {} , max {}".format(sorted_temperatures[0], sorted_temperatures[-1]))
    avg = avg_count / len(sorted_temperatures)
    print("2. avg {}".format(avg))
    count_lessthan_avg = 0
    hot_cool_point = 30
    hot = 0
    cool = 0
    for mew in sorted_temperatures:
        if mew < avg:
            count_lessthan_avg = count_lessthan_avg + 1
        if mew < hot_cool_point:
            cool = cool + 1
        else:
            hot = hot + 1

    print("3. count less than avg {}".format(count_lessthan_avg))
    print("4. hots {} cool {}".format(hot, cool))
    print("5. sorted  {}".format(sorted_temperatures))
    print("6. median {}".format(sorted_temperatures[int(len(sorted_temperatures) / 2)]))


employees = {
    "E001": {"name": "Ali Ahmadi", "dept": "IT", "salary": 5000, "years": 3},
    "E002": {"name": "Sara Hoseini", "dept": "Finance", "salary": 6200, "years": 7},
    "E003": {"name": "Reza Karimi", "dept": "IT", "salary": 4800, "years": 2},
    "E004": {"name": "Mina Tehrani", "dept": "HR", "salary": 5500, "years": 5},
    "E005": {"name": "Kamran Jali", "dept": "Finance", "salary": 7000, "years": 9},
}


# 1. Find the employee with the highest salary
# 2. Group employees by department: {"IT": [...], "Finance": [...], "HR": [...]}
# 3. Calculate average salary per department
# 4. Give a10% raise to employees with more than 5 years of experience
# 5. Create a new dict: {employee_id: name} for all employees
# 6. Find departments that have more than one employee
def func_5():
    print(
        "///////////////////////////////////////// " + inspect.currentframe().f_code.co_name + " /////////////////////////////////////////")
    employees = {
        "E001": {"name": "Ali Ahmadi", "dept": "IT", "salary": 5000, "years": 3},
        "E002": {"name": "Sara Hoseini", "dept": "Finance", "salary": 6200, "years": 7},
        "E003": {"name": "Reza Karimi", "dept": "IT", "salary": 4800, "years": 2},
        "E004": {"name": "Mina Tehrani", "dept": "HR", "salary": 5500, "years": 5},
        "E005": {"name": "Kamran Jali", "dept": "Finance", "salary": 7000, "years": 9},
    }
    max = 0
    max_index = 0
    count = 0
    for employee_num in employees.keys():
        if employees[employee_num]["salary"] > max:
            max = employees[employee_num]["salary"]
            max_index = employee_num
        count = count + int(employees[employee_num]["salary"])
    print(f"1. {employees[max_index]}")
    it = []
    fiance = []
    hr = []
    for employee_num in employees.keys():
        if employees[employee_num]["dept"] == "IT":
            it.append(employees[employee_num]["name"])
        elif employees[employee_num]["dept"] == "Finance":
            fiance.append(employees[employee_num]["name"])
        else:
            hr.append(employees[employee_num]["name"])
    print(f"2. IT :{it}, Fiance :{fiance}, HR :{hr}")
    print("3. avg {}".format(count / len(employees)))
    print("4. Salaries after raise:")
    for employee_id, employee in employees.items():
        if employee["years"] > 5:
            employee["salary"] *= 1.10
        print(f"   {employee_id}: {employee['salary']:.2f}")
    employee_names = {}

    for employee_id, employee in employees.items():
        employee_names[employee_id] = employee["name"]
    print(f"5. {employee_names}")
    department_count = {}
    for employee in employees.values():
        dept = employee["dept"]
        if dept not in department_count:
            department_count[dept] = 0
        department_count[dept] += 1
    multiple_employee_departments = []
    for dept, count in department_count.items():
        if count > 1:
            multiple_employee_departments.append(dept)
    print(f"6. Dept  more than 1 em: " f"{multiple_employee_departments}")


transactions = [
    ("T001", "Ali", "purchase", 250),
    ("T002", "Sara", "refund", 100),
    ("T003", "Ali", "purchase", 350),
    ("T004", "Reza", "purchase", 180),
    ("T005", "Sara", "purchase", 420),
    ("T006", "Ali", "refund", 50),
]


# Tuples are immutable — good for records you shouldn't accidentally modify
# 1. Unpack each tuple and print: "Ali made a purchase of $250"
# 2. Find all unique customers using a set
# 3. Find total spent by each customer (purchases - refunds)
# 4. Find customers who made both purchases AND refunds (set intersection logic)
# 5. Why would using tuples here be safer than lists?
#    (Write your answer as a comment)
def func_6():
    print(
        "///////////////////////////////////////// " + inspect.currentframe().f_code.co_name + " /////////////////////////////////////////")
    transactions = [
        ("T001", "Ali", "purchase", 250),
        ("T002", "Sara", "refund", 100),
        ("T003", "Ali", "purchase", 350),
        ("T004", "Reza", "purchase", 180),
        ("T005", "Sara", "purchase", 420),
        ("T006", "Ali", "refund", 50),
    ]
    customers = set()
    print("1.")
    for code_name, name, kind, count in transactions:
        print(f"  {name} made a {kind} of {count} ")
        customers.add(name)
    print(f"2. {customers}")
    total_spent = {}
    for name_unic in customers:
        total_spent[name_unic] = 0
        for code_name, name, kind, count in transactions:
            if name == name_unic:
                if kind == "purchase":
                    total_spent[name_unic] += count
                elif kind == "refund":
                    total_spent[name_unic] -= count
    print(f"3. {total_spent}")
    purchases = set()
    refunds = set()

    for code_name, name, kind, count in transactions:
        if kind == "purchase":
            purchases.add(name)
        elif kind == "refund":
            refunds.add(name)

    both = purchases & refunds

    print(f"4. {both}")


# benazaram bekhater iterat va packing va unpacking khob un haminjoor hefz eteleaat


# Step 1: Create a file called "log.txt" with this content (write it with Python):
log_entries = """2024-01-15 09:00:01 INFOUser'ali' logged in
2024-01-15 09:05:23 ERROR Database connection timeout
2024-01-15 09:06:01 INFO  Connection restored
2024-01-15 09:10:44 WARNING Memory usage at 85%
2024-01-15 09:15:02 ERROR Null pointer in module auth.py line 42
2024-01-15 09:20:00 INFO  Backup completed successfully
"""


# Step 2: Read the file back and:
# 1. Count lines for each level: INFO, WARNING, ERROR
# 2. Extract only ERROR lines and save to "errors.txt"
# 3. Print a summary: {"INFO": 3, "WARNING": 1, "ERROR": 2}

# Use'with open(...)' — always preferred over open() without context manager
def func_7():
    print(
        "///////////////////////////////////////// " + inspect.currentframe().f_code.co_name + " /////////////////////////////////////////")

    log_entries = """2024-01-15 09:00:01 INFOUser'ali' logged in
    2024-01-15 09:05:23 ERROR Database connection timeout
    2024-01-15 09:06:01 INFO  Connection restored
    2024-01-15 09:10:44 WARNING Memory usage at 85%
    2024-01-15 09:15:02 ERROR Null pointer in module auth.py line 42
    2024-01-15 09:20:00 INFO  Backup completed successfully
    """
    nums_of_levels = {"INFO": 0, "ERROR": 0, "WARNING": 0}
    try:
        with open("log.txt", "w") as file:
            file.write(log_entries)
            print("first step done")
    except FileNotFoundError:
        print("first file not found")
    try:
        with open("log.txt", "r") as file:
            print("second step done")
            for line in file:
                if "INFO" in line.strip().split():
                    nums_of_levels["INFO"] += 1
                elif "ERROR" in line.split():
                    nums_of_levels["ERROR"] += 1
                elif "WARNING" in line.split():
                    nums_of_levels["WARNING"] += 1
            print(f"1. {nums_of_levels}")
            file.seek(0)
            with open("ERROR.txt", "a") as Erorr_file:
                for line in file:
                    if "ERROR" in line.strip().split():
                        Erorr_file.write(line)
                print("2. done")
            print(f"3. {nums_of_levels}")
    except FileNotFoundError:
        print("second file not found")


# Step 1: Create "users.json" with this content (save it manually or write with Python)
# [
#   {"id": 1, "name": "Ali Ahmadi","age": 28, "email": "ali@test.com","active": true,"score": 85},
#   {"id": 2, "name": "Sara Hoseini",  "age": 17, "email": "sara@test.com", "active": true,  "score": 92},
#   {"id": 3, "name": "Reza Karimi",   "age": 35, "email": "bad-email","active": false, "score": 70},
#   {"id": 4, "name": "Mina Tehrani",  "age": 29, "email": "mina@test.com", "active": true,  "score": null},
#   {"id": 5, "name": "Kamran Jalili", "age": 22, "email": "kam@test.com",  "active": true,  "score": 88}
# ]


def func_8():
    print(
        "///////////////////////////////////////// " + inspect.currentframe().f_code.co_name + " /////////////////////////////////////////")

    users = [
        {
            "id": 1, "name": "Ali Ahmadi", "age": 28, "email": "ali@test.com", "active": True, "score": 85},
        {
            "id": 2, "name": "Sara Hoseini", "age": 17, "email": "sara@test.com", "active": True, "score": 92},
        {
            "id": 3, "name": "Reza Karimi", "age": 35, "email": "bad-email", "active": False, "score": 70},
        {
            "id": 4, "name": "Mina Tehrani", "age": 29, "email": "mina@test.com", "active": True, "score": None},
        {
            "id": 5, "name": "Kamran Jalili", "age": 22, "email": "kam@test.com", "active": True, "score": 88}
    ]
    with open("users.json", "w", encoding="utf-8") as file:
        json.dump(users, file, indent=2)
    print("users.json created!")
    with open("users.json", "r") as f:
        users = json.load(f)
    print("1.{}".format(users))
    valid_users = []
    for user in users:
        if user["active"] == True and user["score"] is not None and user["age"] >= 18:
            valid_users.append(user)
    print("2.{}".format(valid_users))
    with open("vali_users.json", "w", encoding="utf-8") as file:
        json.dump(valid_users, file, indent=2, ensure_ascii=False)


def func_9():
    print(
        "///////////////////////////////////////// " + inspect.currentframe().f_code.co_name + " /////////////////////////////////////////")
    orders_saving = [
        {
            "order_id": "A001",
            "customer": {"name": "Ali", "city": "Tehran"},
            "items": [
                {"product": "Laptop", "qty": 1, "price": 5000},
                {"product": "Mouse", "qty": 2, "price": 200}
            ],
            "status": "delivered"
        },
        {
            "order_id": "A002",
            "customer": {"name": "Sara", "city": "Mashhad"},
            "items": [
                {"product": "Keyboard", "qty": 1, "price": 500}
            ],
            "status": "pending"
        },
        {
            "order_id": "A003",
            "customer": {"name": "Reza", "city": "Tehran"},
            "items": [
                {"product": "Monitor", "qty": 2, "price": 3000},
                {"product": "Mouse", "qty": 1, "price": 200}
            ],
            "status": "delivered"
        }
    ]
    with open("orders.json", "w", encoding="utf-8") as file:
        json.dump(orders_saving, file, indent=2, ensure_ascii=False)
    orders_price = {}
    delivered = []
    tehranion = []
    behtarion = defaultdict(int)
    order_count = 0
    delivered_count = 0
    with open("orders.json", "r", encoding="utf-8") as file:
        orders = json.load(file)
        for order in orders:
            temp = 0
            for item in order["items"]:
                temp += item["price"] * item["qty"]
            orders_price[order["order_id"]] = temp
        print("1.{}".format(orders_price))
        for order in orders:
            if order["status"] == "delivered":
                delivered.append(order)
        print("2.{}".format(delivered))
        for order in orders:
            if order["customer"]["city"] == "Tehran":
                tehranion.append(order["customer"])
        print("3.{}".format(tehranion))
        for order in orders:
            for item in order["items"]:
                behtarion[item["product"]] += item["qty"]
        max_products = max(behtarion)
        print("4. best product{}".format(max_products))
        order_count = len(orders)
        delivered_count = len(delivered)
    summary = {}
    summary["order_count"] = order_count
    summary["delivered_count"] = delivered_count
    summary["best_product"] = max_products
    summary["orders"] = orders_price
    with open("summary.json", "w", encoding="utf-8") as file:
        json.dump(summary, file, indent=2, ensure_ascii=False)
        print("5. summary.json created!")


def func_10():
    print(
        "///////////////////////////////////////// " + inspect.currentframe().f_code.co_name + " /////////////////////////////////////////")

    wb = openpyxl.load_workbook("students.xlsx")
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    students = []
    best_per_city = {}
    students_avg_grade = {}
    result = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        student = dict(zip(headers, row))
        students.append(student)
    for student in students:
        avg_score = 0
        avg_score = (student["Math"] + student["Science"] + student["English"]) / 3
        city = student["City"]
        students_avg_grade[student["Name"]] = avg_score
        if city not in best_per_city or avg_score > best_per_city[city][1]:
            best_per_city[city] = (student, avg_score)
    result.append(students_avg_grade)
    result.append(best_per_city)
    print("1. 2. 3. 4. {}".format(result))
    with open("result.json", "w", encoding="utf-8") as file:
        json.dump(result, file, indent=2, ensure_ascii=False)
    ws2 = wb.create_sheet(title="Results")
    ws2.append(["Name", "City", "Average"])
    for s in students:
        avg = round((s["Math"] + s["Science"] + s["English"]) / 3, 1)
        ws2.append([s["Name"], s["City"], avg])

    wb.save("students.xlsx")
    print("New sheet'Results' added to students.xlsx")


def func_11():
    print(
        "///////////////////////////////////////// " + inspect.currentframe().f_code.co_name + " /////////////////////////////////////////")

    with open("users.json", "r") as f:
        users = json.load(f)
    print("1. ")
    for user in users:
        print(f"{user['name']} => {user['score']}")
    for user in users:
        user["city"] = random.choice(["Tehran", "Mashhad", "Esfahan", "Tehran", "Yazd"])
    valid_cities = set()
    for user in users:
        valid_cities.add(user["city"])
    print(valid_cities)
    while True:
        input_city = input("Please enter your city: ")

        if input_city == "exit":
            break
        elif input_city not in valid_cities:
            print("Invalid city")
        else:
            for user in users:
                if user["city"] == input_city:
                    print(f"{user['name']} ")
    active_users = [user for user in users if user["active"] == True]
    high_scores = [user for user in users if user["score"] is not None and int(user["score"]) > 85]
    print("2. ")
    print(high_scores)
    print("3. ")
    print(active_users)
    matrix = [[1, 2, 3], [4, 5, 6], [7, 8, 9]]
    new_matrix = [num for line in matrix for num in line]
    print(f"6.{new_matrix}")


def func_12():
    print(
        "///////////////////////////////////////// " + inspect.currentframe().f_code.co_name + " /////////////////////////////////////////")
    wb = openpyxl.load_workbook("students.xlsx")
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        math = row[2]
        science = row[3]
        english = row[4]
        average = (math + science + english) / 3
        if average >= 80:
              print(f"{name}: {average:.2f}")
    grades = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        math = row[2]
        science = row[3]
        english = row[4]
        average = (math + science + english) / 3
        if average >= 90:
            grade = "A"
        elif average >= 80:
            grade = "B"
        elif average >= 70:
            grade = "C"
        elif average >= 60:
            grade = "D"
        else:
            grade = "F"
        grades[name] = grade
    print(f"2. {grades}")
    grade_count = {}
    for grade in grades.values():
        if grade not in grade_count:
            grade_count[grade] = 0
        grade_count[grade] += 1
    print(f"3. {grade_count}")
    students_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]
        math = row[2]
        science = row[3]
        english = row[4]
        average = (math + science + english) / 3
        students_data.append((name, average))
    overall_average = sum(
        average for name, average in students_data
    ) / len(students_data)
    below_average = [
        name
        for name, average in students_data
        if average < overall_average
    ]
    print(f"4. {below_average}")
    students_sorted = sorted(
        students_data,
        key=lambda student: student[1],
        reverse=True
    )
    students = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        students.append({
            "Name": row[0],
            "Age": row[1],
            "Math": row[2],
            "Science": row[3],
            "English": row[4],
            "City": row[5]
        })
    students_sorted = sorted(
        students,
        key=lambda s: (s["Math"] + s["Science"] + s["English"]) / 3,
        reverse=True
    )
    print(f"5. {students_sorted}")
# ///////////////////////////////////////////////////////////////////////////////////////////////////
def load_json(filepath):
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"File not found: {filepath}")
        return None
    except json.JSONDecodeError:
        print(f"Invalid JSON: {filepath}")
        return None
def save_json(data, filepath, indent=2):
    try:
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(
                data,
                f,
                indent=indent,
                ensure_ascii=False
            )
        return True
    except OSError as e:
        print(f"Error saving file: {e}")
        return False
def load_csv(filepath):
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            return [dict(row) for row in reader]
    except FileNotFoundError:
        print(f"File not found: {filepath}")
        return None
def filter_records(records, field, value):
    return [
        record
        for record in records
        if record.get(field) == value
    ]
def get_field_values(records, field):
    return [
        record[field]
        for record in records
        if record.get(field) is not None
    ]
def summarize(records, numeric_field):
    values = get_field_values(records, numeric_field)
    if not values:
        return {
            "count": 0,
            "min": None,
            "max": None,
            "average": None
        }
    return {
        "count": len(values),
        "min": min(values),
        "max": max(values),
        "average": sum(values) / len(values)
    }


# # Part A: *args and **kwargs
# def generate_report(*records, **options):
#     """
#     records : any number of dicts
#     options:- title(str): report title
#       - fields  (list) : which fields to display
#       - sort_by (str)  : field name to sort by
#       - limit   (int)  : max number of records to show
#
#     Example usage:
#       generate_report(*users,title="Top Users",
#                       fields=["name", "score"],
#                       sort_by="score",
#                       limit=3)
#     """
#     pass
#
#
# # Part B: Lambda, map, filter
# import json
#
# with open("users.json") as f:
#     users = json.load(f)
#
# # 1. Use map() to add a 'grade' key to each user (A/B/C/D based on score)
# # 2. Use filter() to keep only active users with non-null scores
# # 3. Use sorted() with lambda to sort by score descending
# # 4. Chain them: filter → sort → add grade → save to "graded_users.json"
#
# # Part C: Scope question
# x = 10
#
#
# def outer():
#     x = 20
#
#     def inner():
#         x = 30
#         print("inner x:", x)  # what prints here?
#
#     inner()
#     print("outer x:", x)  # what prints here?
#
#
# outer()
# print("global x:", x)  # what prints here?
#
#
# # Predict the output, then run it. Were you right?

def generate_report(*records, **options):
    title = options.get("title", "Report")
    fields = options.get("fields")
    sort_by = options.get("sort_by")
    limit = options.get("limit")
    records = list(records)

    if sort_by:
        records.sort(
            key=lambda record: record.get(sort_by, 0),
            reverse=True
        )
    if limit is not None:
        records = records[:limit]
    print("=" * 40)
    print(title)
    print("=" * 40)
    for record in records:
        if fields:
            result = {}
            for field in fields:
                result[field] = record.get(field)
            print(result)
        else:
            print(record)

def get_grade(score):
    if score is None:
        return None

    if score >= 90:
        return "A"
    elif score >= 80:
        return "B"
    elif score >= 70:
        return "C"
    elif score >= 60:
        return "D"
    else:
        return "F"
#//////////////////////////////////////////////////////////////////////////////////////////////////////
def get_difficulty():
    while True:
        difficulty = input("Enter difficulty:(1/2/3) ")
        if difficulty == "1":
            return 10
        elif difficulty == "2":
            return 7
        elif difficulty == "3":
            return 5
        else :
            print("Invalid difficulty please enter 1 or of 2 or 3")
def get_guess(low , high):
    while True:
        guess = input("Enter guess:")
        try:
            guess=int(guess)
            if guess < low or guess > high:
              print("Invalid guess please enter lower or higher")
            else :
                return guess
        except ValueError:
            print("Invalid guess please enter numeric guess")
def check_guess(guess, secret):
    if guess == secret:
        print("Correct!")
        return True
    elif guess < secret:
        print("too_low")
        return False
    else :
        print("too_high")
        return False
def calculate_score(guesses_used, max_guesses):
    return (max_guesses -  guesses_used) * (1000 / max_guesses)
def load_highscores(filepath="highscores.json"):
    return load_json(filepath)
def save_highscore(name, score, difficulty, filepath="highscores.json"):
    new_highscore = {}
    new_highscore["name"] = name
    new_highscore["score"] = score
    new_highscore["difficulty"] = difficulty
    formatted_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    new_highscore["Time"] = formatted_time
    high_score=(load_highscores(filepath))
    if high_score is not None:
        high_score.append(new_highscore)
        sorted_high_scores=sorted(high_score, key=lambda h: h["score"], reverse=True)
        top_5 = sorted_high_scores[:5]
        save_json(top_5, filepath)
    else :
        save_json([new_highscore], filepath)

def show_highscores(filepath="highscores.json"):
     top_5=load_json(filepath)
     i =1
     for h in top_5:
         print(f"{i}. is {h['name']} by {h['score']} in {h['difficulty']} in {h['Time']}")
         i += 1

def play_game():

    while True:
        hoom=input("Would you like to do ? (play/show high scores( by show )/ exit)")
        if hoom == "play":
            name = input("ur name: ")
            con = True
            while con:
                difficulty = get_difficulty()
                secret = random.randint(1, 100)
                print(secret)
                high = 100
                low = 0
                i = 0
                while difficulty > i:
                    guess = get_guess(low, high)
                    if check_guess(guess, secret):
                        break
                    else:
                        i+=1
                score = calculate_score(i,difficulty)
                print(f"hey {name} ur score is:", score)
                if difficulty == 10 :
                    dif_massage= "easy"
                elif difficulty == 7 :
                    dif_massage= "hard"
                elif difficulty == 5 :
                    dif_massage= "too_hard"
                save_highscore(name, score,dif_massage)
                show_highscores()
                while True:
                    a=input("Would you like to play again? (y/n)")
                    if a == "n":
                        con = False
                        break
                    elif a == "y":
                        pass
                    else :
                        print("invalid input")
        elif hoom == "show":
            show_highscores()
        elif hoom == "exit" :
            break
        else :
            print("Invalid input")

class DataFile:

    def __init__(self, filepath):
        self.filepath = filepath
        self.data = []

    def load(self):
        """Must be implemented by subclasses."""
        raise NotImplementedError("Subclass must implement load()")

    def save(self, output_path, data=None):
        """Must be implemented by subclasses."""

    def filter(self, field, value):
        return [r for r in self.data if r.get(field) == value]

    def get_values(self, field):
        return [r[field] for r in self.data if r.get(field) is not None]

    def summary(self):
        if not self.data:
            print("No data loaded.")
            return
        print(f"File: {self.filepath}")
        print(f"Records: {len(self.data)}")
        print(f"Fields : {list(self.data[0].keys())}")
    def __len__(self):
        return len(self.data)
    def __repr__(self):
        return f"{self.__class__.__name__}('{self.filepath}', {len(self.data)} records)"


class JSONFile(DataFile):
    def load(self):
        with open(self.filepath, "r", encoding="utf-8") as f:
            self.data = json.load(f)
        print(f"Loaded {len(self.data)} records from {self.filepath}")
    def save(self, output_path, data=None):
        to_save = data if data is not None else self.data
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(to_save, f, indent=2, ensure_ascii=False)
        print(f"Saved {len(to_save)} records to {output_path}")


class CSVFile(DataFile):
    def load(self):
        with open(self.filepath, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            self.data = [dict(row) for row in reader]
        print(f"Loaded {len(self.data)} records from {self.filepath}")
    def save(self, output_path, data=None):
        to_save = data if data is not None else self.data
        if not to_save:
            return
        with open(output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=to_save[0].keys()
            )
            writer.writeheader()
            writer.writerows(to_save)
        print(f"Saved {len(to_save)} records to {output_path}")


class ExcelFile(DataFile):
    def load(self):
        import openpyxl
        wb = openpyxl.load_workbook(self.filepath)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        self.data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            self.data.append(dict(zip(headers, row)))
        print(f"Loaded {len(self.data)} records from {self.filepath}")
    def save(self, output_path, data=None):
        import openpyxl
        to_save = data if data is not None else self.data
        if not to_save:
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = list(to_save[0].keys())
        ws.append(headers)
        for record in to_save:
            ws.append([record.get(header) for header in headers])
        wb.save(output_path)
        print(f"Saved {len(to_save)} records to {output_path}")

def fun_16():
    jf = JSONFile("data.json")
    jf.load()
    print(jf)
    print(len(jf))
    jf.summary()
    print(jf.filter("name", "Ali"))
    print(jf.get_values("name"))
    jf.save("output.json")
    cf = CSVFile("data.csv")
    cf.load()
    print(cf)
    print(len(cf))
    cf.summary()
    print(cf.filter("name", "Ali"))
    print(cf.get_values("name"))
    cf.save("output.csv")
    ef = ExcelFile("data.xlsx")
    ef.load()
    print(ef)
    print(len(ef))
    ef.summary()
    print(ef.filter("name", "Ali"))
    print(ef.get_values("name"))
    ef.save("output.xlsx")

if __name__ == "__main__":
    # func_1()
    # func_2()
    # func_3()
    # func_4()
    # func_5()
    # func_6()
    # func_7()
    # func_8()
    # func_9()
    # func_10()
    # func_11()
    play_game()
    fun_16()